import json
import re
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUTPUT_JSON_DIR = Path("output/json")
OUTPUT_REPORTS_DIR = Path("output/reports")

ORDERED_HEADERS = [
    "Cliente",
    "Código JS",
    "Referência Cliente",
    "Responsável Atual",
    "Status",
    "Modal",
    "Tipo Transporte",
    "Regime",
    "Embarcação/Plataforma",
    "Incoterm",
    "ETA",
    "Data de Embarque",
    "Data de Chegada",
    "Data de Registro da DI",
    "Numero da DI",
    "Data da CI",
    "Canal",
    "House AWB/BL",
    "Master AWB/BL",
    "Pack List",
    "Commercial Invoice",
    "Proforma Invoice",
    "País de Origem",
    "Local de Embarque",
    "URF Entrada no País",
    "Peso Líquido",
    "Peso Bruto",
    "Volume",
    "Valor Declarado",
    "Moeda",
    "Valor da Taxa",
    "Frete Collect",
    "Frete Prepaid",
    "Frete Nacional",
    "Moeda Frete",
    "Recinto Aduaneiro",
    "Armazém",
]


LABEL_MARKERS = [
    "Cliente",
    "Código JS",
    "Referência Cliente",
    "Responsável Atual",
    "Status Processo",
    "Vinculado ao Processo",
    "Dados Gerais",
    "Documentos",
    "Informações Adicionais",
    "Ocorrência",
    "Demonstrativo de Despesas",
    "Tipo Processo",
    "Nr. DI",
    "Numero da DI",
    "Número da DI",
    "Contrato",
    "Dt.Registro DI",
    "Data Registro DI",
    "Tipo Declaração",
    "Taxa Cambio",
    "Taxa Câmbio",
    "Valor Declarado",
    "Data Moeda Declarada",
    "Moeda",
    "Valor da Taxa",
    "Conhecimento / Transporte",
    "Tipo Transporte",
    "Conhecimento",
    "Data Embarque",
    "Data de Embarque",
    "Local Embarque",
    "Local de Embarque",
    "Doc.Chegada de Carga",
    "Identificação Master",
    "Identificação Chegada Carga",
    "Pack List",
    "Nome Transportador",
    "País Transportador",
    "Identificação",
    "Carga País Procedência da Carga",
    "País Procedência da Carga",
    "Pais Procedencia da Carga",
    "País de Origem",
    "Pais de Origem",
    "Moeda Frete",
    "URF Entrada no Pais",
    "URF Entrada no País",
    "Frete Prepaid",
    "Frete Collect",
    "Frete Nacional",
    "Peso Liquido",
    "Peso Líquido",
    "Peso Bruto",
    "Incoterm",
    "Moeda Seguro",
    "Taxa Siscomex",
    "Valor Seguro",
    "Volumes Embalagem",
    "Volumes",
    "Volume",
    "Quantidade",
    "Chegada",
    "Data Chegada Carga",
    "Data de Chegada",
    "URF de Despacho",
    "Recinto Aduaneiro",
    "Setor de Armazenamento",
    "Armazém",
    "Armazem",
    "Invoice(s)",
    "Nr.Invoice",
    "Exportador",
]


def latest_consolidated_json() -> Path:
    files = sorted(OUTPUT_JSON_DIR.glob("followup_consolidado_*.json"))
    if not files:
        raise FileNotFoundError(
            "Nenhum followup_consolidado_*.json encontrado em output/json. "
            "Rode primeiro o main.py com os processos."
        )
    return files[-1]


def clean_simple_value(value):
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    text = re.sub(r"[ \t]+", " ", text)
    return text


def to_single_line(value):
    text = clean_simple_value(value)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def cut_at_markers(value, markers):
    text = clean_simple_value(value)
    if not text:
        return ""

    cut_positions = []
    for marker in markers:
        if not marker:
            continue

        pos = text.find(marker)
        if pos >= 0:
            cut_positions.append(pos)

        # fallback case-insensitive, porque às vezes o texto vem com variação
        lowered_pos = text.lower().find(str(marker).lower())
        if lowered_pos >= 0:
            cut_positions.append(lowered_pos)

    if cut_positions:
        first_pos = min(cut_positions)
        if first_pos == 0:
            return ""
        text = text[:first_pos]

    return to_single_line(text).strip(" .:-")


NOISE_EXACT_VALUES = {
    "Taxa Cambio",
    "Taxa Câmbio",
    "Conhecimento / Transporte",
    "Carga",
    "Volumes",
    "Chegada",
    "Invoice(s)",
    "Ocorrências",
    "Pagamentos na Conta do Cliente",
    "Nova Ocorrência",
    "Novo Documento",
    "Nr.Invoice",
    "Valor na Moeda",
    "Exportador",
}


def is_noise_value(value):
    text = to_single_line(value).strip(" .:-")
    if not text:
        return True

    lowered = text.lower()

    if "has_processo_word" in lowered or "has_resultado_consulta" in lowered:
        return True

    if "has_usuario_word" in lowered or "has_senha_word" in lowered:
        return True

    label_values = {to_single_line(label).lower() for label in LABEL_MARKERS}
    label_values.update(to_single_line(label).lower() for label in NOISE_EXACT_VALUES)

    if lowered in label_values:
        return True

    if lowered.startswith("invoice(s)") and ("nr.invoice" in lowered or "exportador" in lowered):
        return True

    return False


def sanitize_value(value):
    text = to_single_line(value)
    if is_noise_value(text):
        return ""
    return text


def clean_field(value, *markers):
    return sanitize_value(cut_at_markers(value, list(markers)))


def joined_dados_text(dados):
    if not isinstance(dados, dict):
        return ""

    # Não junte todos os valores do dict. O dict tem "tables" e "indicators";
    # quando isso vira texto, o export acaba pescando metadados como se fossem campos.
    return to_single_line(dados.get("preview", ""))


def extract_label_value(text, label):
    """
    Extrai valores de blocos grandes no formato:
    Label: valor Próximo Label: outro valor

    Isso resolve casos em que o site devolve um blocão inteiro em um campo,
    principalmente em processos ADMINISTRATIVO / COMERCIAL.
    """
    text = to_single_line(text)
    if not text:
        return ""

    match = re.search(rf"{re.escape(label)}\s*:\s*", text, flags=re.IGNORECASE)
    if not match:
        return ""

    start = match.end()
    stop_positions = []

    for marker in LABEL_MARKERS:
        if marker.lower() == label.lower():
            continue

        next_match = re.search(
            rf"\b{re.escape(marker)}\s*:",
            text[start:],
            flags=re.IGNORECASE,
        )

        if next_match:
            stop_positions.append(start + next_match.start())

    end = min(stop_positions) if stop_positions else len(text)
    value = text[start:end]

    return sanitize_value(clean_simple_value(value).strip(" .:-"))


def first_meaningful_line(value):
    text = clean_simple_value(value)
    if not text:
        return ""

    stop_words = {
        "referência",
        "responsável",
        "status",
        "vinculado",
        "dados gerais",
        "documentos",
        "informações adicionais",
        "ocorrência",
        "demonstrativo de despesas",
    }

    for raw_line in str(value).splitlines():
        line = raw_line.strip(" .:\t\r\n")
        if not line:
            continue

        lowered = line.lower()
        if lowered in stop_words:
            break

        if any(lowered.startswith(sw) for sw in stop_words):
            break

        if line not in {"-", ":"}:
            return line

    return to_single_line(text)


def normalize_modal(tipo_processo, tipo_transporte):
    tipo_processo_limpo = clean_field(tipo_processo, 
        "Nr. DI", 
        "Contrato", 
        "Dt.Registro DI", 
        "Tipo Declaração", 
        "Taxa Cambio", 
        "Valor Declarado",
        "Conhecimento/Transporte",
        "Tipo Transporte",
    )

    tipo_transporte_limpo = clean_field(tipo_transporte,
        "Conhecimento",
        "Data Embarque",
        "Local Embarque",
        "Doc.Chegada de Carga",
    )

    def modal_from_text(value):
        text = to_single_line(value).upper()

        if "ADMINISTRATIVO" in text:
            return "ADMINISTRATIVO"
        
        if "COMERCIAL" in text:
            return "COMERCIAL"
        
        if "AERE" in text or "AÉRE" in text:
            return "AÉREO"
        
        if "MARIT" in text or "MARÍT" in text:
            return "MARÍTIMO"
        
        if "RODOV" in text:
            return "RODOVIÁRIO"
    
        return ""
    
    modal = modal_from_text(tipo_processo_limpo)
    if modal:
        return modal
        
    modal = modal_from_text(tipo_transporte_limpo)
    if modal:
        return modal
        
    return tipo_processo_limpo or tipo_transporte_limpo


def get_section_tables(section):
    if not isinstance(section, dict):
        return []
    return section.get("tables", []) or []


def extract_doc_rows(section):
    results = []

    for table in get_section_tables(section):
        rows = table.get("rows", [])
        if len(rows) < 2:
            continue

        header = [to_single_line(x).lower() for x in rows[0]]
        joined = " | ".join(header)

        if "tipo documento" not in joined or "nr.documento" not in joined:
            continue

        for row in rows[1:]:
            if not row:
                continue

            first = to_single_line(row[0]).lower()
            if first.startswith("total"):
                continue

            results.append({
                "tipo_documento": to_single_line(row[0]) if len(row) > 0 else "",
                "nr_documento": to_single_line(row[1]) if len(row) > 1 else "",
                "data_entrega": to_single_line(row[2]) if len(row) > 2 else "",
                "data_validade": to_single_line(row[3]) if len(row) > 3 else "",
            })

    return results


def find_doc(rows, *wanted_types):
    wanted = {w.upper() for w in wanted_types}
    for row in rows:
        if to_single_line(row.get("tipo_documento")).upper() in wanted:
            return row
    return None


def first_non_empty(*values):
    for value in values:
        value = sanitize_value(value)
        if value:
            return value
    return ""


def extract_from_preview(preview, label):
    text = clean_simple_value(preview)
    if not text:
        return ""

    pattern = rf"{re.escape(label)}\s*=>\s*([^\r\n]+)"
    match = re.search(pattern, text, flags=re.IGNORECASE)
    return sanitize_value(match.group(1)) if match else ""


def parse_data_chegada_from_preview(preview):
    text = clean_simple_value(preview)
    match = re.search(
        r"Data Chegada Carga:\s*([0-9]{2}/[0-9]{2}/[0-9]{4})",
        text,
        flags=re.IGNORECASE,
    )
    return sanitize_value(match.group(1)) if match else ""


def build_row(process_data):
    resumo = process_data.get("resumo", {}) or {}
    dados = process_data.get("dados_gerais", {}) or {}
    documentos = process_data.get("documentos", {}) or {}
    adicionais = process_data.get("documentos_adicionais", {}) or {}
    ocorrencia = process_data.get("ocorrencia", {}) or {}

    doc_rows = extract_doc_rows(documentos)
    add_rows = extract_doc_rows(adicionais)
    rows_all = doc_rows + add_rows

    doc_di = find_doc(rows_all, "DI")
    doc_ci = find_doc(rows_all, "CI", "CI (COMPROV. IMPORT)")
    doc_house = find_doc(rows_all, "HAWB/HBL", "HAWB", "HBL")
    doc_master = find_doc(rows_all, "AWB/MAWB", "MAWB", "AWB", "MASTER BL", "MASTER B/L")
    doc_bl = find_doc(rows_all, "B/L", "BL", "BILL OF LADING")
    doc_commercial = find_doc(rows_all, "INVOICE - COMMERCIAL")
    doc_proforma = find_doc(rows_all, "INVOICE - PROFORMA", "PROFORMA")
    doc_ref_cliente = find_doc(rows_all, "REFERENCIA CLIENTE")
    doc_pais_origem = find_doc(rows_all, "PAÍS DE ORIGEM")
    doc_peso_bruto = find_doc(rows_all, "PESO BRUTO")
    doc_volumes = find_doc(rows_all, "VOLUMES")
    doc_data_registro_di = find_doc(rows_all, "DATA REGISTRO DI")
    doc_canal = find_doc(rows_all, "CANAL")
    doc_eta = find_doc(rows_all, "PREVISÃO DE CHEGADA", "ETA")
    doc_pack_list = find_doc(rows_all, "ROMANEIO CARGA / PACKING LIST", "PACK LIST")

    occ_preview = ocorrencia.get("preview", "")
    dados_preview = dados.get("preview", "")
    dados_blob = joined_dados_text(dados)

    modal = normalize_modal(dados.get("tipo_processo"), dados.get("tipo_transporte"))
    is_admin_comercial = modal in {"ADMINISTRATIVO", "COMERCIAL"}

    cliente = first_meaningful_line(resumo.get("cliente"))
    referencia_resumo = first_meaningful_line(resumo.get("referencia"))

    referencia_cliente = first_non_empty(
        doc_ref_cliente["nr_documento"] if doc_ref_cliente else "",
        extract_from_preview(occ_preview, "REFERENCIA CLIENTE"),
        referencia_resumo,
    )

    responsavel_atual = clean_field(
        resumo.get("responsavel_atual"),
        "Status Processo",
        "Vinculado ao Processo",
        "Dados Gerais",
        "Documentos",
    )

    status = clean_field(
        resumo.get("status_processo"),
        "Vinculado ao Processo",
        "Dados Gerais",
        "Documentos",
        "Informações Adicionais",
    )

    tipo_transporte_limpo = "" if is_admin_comercial else clean_field(
        dados.get("tipo_transporte"),
        "Data Embarque",
        "Conhecimento",
        "Local Embarque",
    )

    regime_limpo = first_non_empty(
        clean_field(
            dados.get("tipo_declaracao"),
            "Taxa Cambio",
            "Taxa Câmbio",
            "Valor Declarado",
            "Data Moeda Declarada",
        ),
        extract_label_value(dados_blob, "Tipo Declaração"),
    )

    embarcacao_plataforma_limpa = clean_field(
        dados.get("embarcacao_plataforma"),
        "Tipo Processo",
        "Contrato",
    )

    incoterm_limpo = first_non_empty(
        clean_field(
            dados.get("incoterm"),
            "Moeda Seguro",
            "Taxa Siscomex",
            "Valor Seguro",
            "Volumes",
            "Chegada",
        ),
        extract_label_value(dados_blob, "Incoterm"),
    )

    numero_di = first_non_empty(
        clean_field(
            dados.get("nr_di"),
            "Contrato",
            "Dt.Registro DI",
            "Tipo Declaração",
        ),
        extract_label_value(dados_blob, "Nr. DI"),
        doc_di["nr_documento"] if doc_di else "",
    )

    data_registro_di = first_non_empty(
        clean_field(
            dados.get("dt_registro_di"),
            "Tipo Declaração",
            "Taxa Cambio",
            "Taxa Câmbio",
        ),
        extract_label_value(dados_blob, "Dt.Registro DI"),
        extract_label_value(dados_blob, "Data Registro DI"),
        doc_data_registro_di["nr_documento"] if doc_data_registro_di else "",
        doc_data_registro_di["data_entrega"] if doc_data_registro_di else "",
    )

    data_ci = first_non_empty(
        doc_ci["data_entrega"] if doc_ci else "",
        extract_from_preview(occ_preview, "DATA DA CI"),
    )

    canal = first_non_empty(
        doc_canal["nr_documento"] if doc_canal else "",
        extract_from_preview(occ_preview, "CANAL"),
    )

    eta = first_non_empty(
        doc_eta["nr_documento"] if doc_eta else "",
        extract_from_preview(occ_preview, "PREVISÃO DE CHEGADA"),
        extract_from_preview(occ_preview, "ETA"),
    )

    data_embarque_limpa = first_non_empty(
        clean_field(
            dados.get("data_embarque"),
            "Local Embarque",
            "Doc.Chegada de Carga",
        ),
        extract_label_value(dados_blob, "Data Embarque"),
        extract_label_value(dados_blob, "Data de Embarque"),
    )

    data_chegada = first_non_empty(
        clean_field(
            dados.get("data_chegada_carga"),
            "URF de Despacho",
            "Recinto Aduaneiro",
            "Setor de Armazenamento",
        ),
        parse_data_chegada_from_preview(dados_preview),
        extract_label_value(dados_blob, "Data Chegada Carga"),
        extract_label_value(dados_blob, "Data de Chegada"),
    )

    house_awb_bl = first_non_empty(
        doc_house["nr_documento"] if doc_house else "",
    )

    master_awb_bl = first_non_empty(
        doc_bl["nr_documento"] if doc_bl else "",
        doc_master["nr_documento"] if doc_master else "",
    )

    commercial_invoice = first_non_empty(
        doc_commercial["nr_documento"] if doc_commercial else "",
        extract_label_value(dados_blob, "Nr.Invoice"),
    )

    proforma_invoice = first_non_empty(
        doc_proforma["nr_documento"] if doc_proforma else "",
    )

    pais_origem = first_non_empty(
        doc_pais_origem["nr_documento"] if doc_pais_origem else "",
        clean_field(
            dados.get("pais_procedencia_carga"),
            "Moeda Frete",
            "URF Entrada no Pais",
            "URF Entrada no País",
            "Frete Prepaid",
        ),
        extract_label_value(dados_blob, "Carga País Procedência da Carga"),
        extract_label_value(dados_blob, "País Procedência da Carga"),
        extract_label_value(dados_blob, "Pais Procedencia da Carga"),
        extract_label_value(dados_blob, "País de Origem"),
        extract_label_value(dados_blob, "Pais de Origem"),
    )

    local_embarque_limpo = first_non_empty(
        clean_field(
            dados.get("local_embarque"),
            "Doc.Chegada de Carga",
            "Identificação Master",
        ),
        extract_label_value(dados_blob, "Local Embarque"),
        extract_label_value(dados_blob, "Local de Embarque"),
    )

    urf_entrada_limpa = first_non_empty(
        clean_field(
            dados.get("urf_entrada_pais"),
            "Peso Líquido",
            "Peso Liquido",
            "Peso Bruto",
            "Incoterm",
        ),
        extract_label_value(dados_blob, "URF Entrada no País"),
        extract_label_value(dados_blob, "URF Entrada no Pais"),
    )

    peso_liquido_limpo = first_non_empty(
        clean_field(
            dados.get("peso_liquido"),
            "Peso Bruto",
            "Incoterm",
        ),
        extract_label_value(dados_blob, "Peso Líquido"),
        extract_label_value(dados_blob, "Peso Liquido"),
    )

    peso_bruto = first_non_empty(
        doc_peso_bruto["nr_documento"] if doc_peso_bruto else "",
        clean_field(
            dados.get("peso_bruto"),
            "Frete Nacional",
            "Incoterm",
            "Moeda Seguro",
            "Taxa Siscomex",
        ),
        extract_label_value(dados_blob, "Peso Bruto"),
    )

    volume = first_non_empty(
        doc_volumes["nr_documento"] if doc_volumes else "",
        extract_label_value(dados_blob, "Volume"),
        extract_label_value(dados_blob, "Volumes"),
    )

    pack_list = first_non_empty(
        clean_field(
            dados.get("pack_list"),
            "Nome Transportador",
            "País Transportador",
        ),
        extract_label_value(dados_blob, "Pack List"),
        doc_pack_list["nr_documento"] if doc_pack_list else "",
    )

    valor_declarado_limpo = first_non_empty(
        clean_field(
            dados.get("valor_declarado"),
            "Data Moeda Declarada",
            "Moeda",
            "Valor da Taxa",
            "Taxa Cambio",
            "Taxa Câmbio",
        ),
        extract_label_value(dados_blob, "Valor Declarado"),
    )

    moeda_limpa = first_non_empty(
        clean_field(
            dados.get("moeda"),
            "Valor da Taxa",
            "Conhecimento / Transporte",
            "Tipo Transporte",
        ),
        extract_label_value(dados_blob, "Moeda"),
    )

    valor_taxa_limpo = first_non_empty(
        clean_field(
            dados.get("valor_taxa_cambio"),
            "Conhecimento / Transporte",
            "Tipo Transporte",
        ),
        extract_label_value(dados_blob, "Valor da Taxa"),
        extract_label_value(dados_blob, "Taxa Cambio"),
        extract_label_value(dados_blob, "Taxa Câmbio"),
    )

    frete_collect_limpo = first_non_empty(
        clean_field(
            dados.get("frete_collect"),
            "Peso Bruto",
            "Frete Nacional",
        ),
        extract_label_value(dados_blob, "Frete Collect"),
    )

    frete_prepaid_limpo = first_non_empty(
        clean_field(
            dados.get("frete_prepaid"),
            "Peso Líquido",
            "Peso Liquido",
            "Frete Collect",
        ),
        extract_label_value(dados_blob, "Frete Prepaid"),
    )

    frete_nacional_limpo = first_non_empty(
        clean_field(
            dados.get("frete_nacional"),
            "Incoterm",
            "Moeda Seguro",
        ),
        extract_label_value(dados_blob, "Frete Nacional"),
    )

    moeda_frete_limpa = first_non_empty(
        clean_field(
            dados.get("moeda_frete"),
            "URF Entrada no País",
            "URF Entrada no Pais",
            "Frete Prepaid",
        ),
        extract_label_value(dados_blob, "Moeda Frete"),
    )

    recinto_aduaneiro_limpo = first_non_empty(
        clean_field(
            dados.get("recinto_aduaneiro"),
            "Setor de Armazenamento",
            "Armazém",
            "Armazem",
            "Invoice(s)",
        ),
        extract_from_preview(dados_preview, "Recinto Aduaneiro"),
        extract_label_value(dados_blob, "Recinto Aduaneiro"),
    )

    armazem_limpo = first_non_empty(
        clean_field(
            dados.get("armazem"),
            "Invoice(s)",
            "Ocorrências",
        ),
        extract_from_preview(dados_preview, "Armazém"),
        extract_label_value(dados_blob, "Armazém"),
        extract_label_value(dados_blob, "Armazem"),
    )

    return {
        "Cliente": cliente,
        "Código JS": process_data.get("process_id", ""),
        "Referência Cliente": referencia_cliente,
        "Responsável Atual": responsavel_atual,
        "Status": status,
        "Modal": modal,
        "Tipo Transporte": tipo_transporte_limpo,
        "Regime": regime_limpo,
        "Embarcação/Plataforma": embarcacao_plataforma_limpa,
        "Incoterm": incoterm_limpo,
        "ETA": eta,
        "Data de Embarque": data_embarque_limpa,
        "Data de Chegada": data_chegada,
        "Data de Registro da DI": data_registro_di,
        "Numero da DI": numero_di,
        "Data da CI": data_ci,
        "Canal": canal,
        "House AWB/BL": house_awb_bl,
        "Master AWB/BL": master_awb_bl,
        "Pack List": pack_list,
        "Commercial Invoice": commercial_invoice,
        "Proforma Invoice": proforma_invoice,
        "País de Origem": pais_origem,
        "Local de Embarque": local_embarque_limpo,
        "URF Entrada no País": urf_entrada_limpa,
        "Peso Líquido": peso_liquido_limpo,
        "Peso Bruto": peso_bruto,
        "Volume": volume,
        "Valor Declarado": valor_declarado_limpo,
        "Moeda": moeda_limpa,
        "Valor da Taxa": valor_taxa_limpo,
        "Frete Collect": frete_collect_limpo,
        "Frete Prepaid": frete_prepaid_limpo,
        "Frete Nacional": frete_nacional_limpo,
        "Moeda Frete": moeda_frete_limpa,
        "Recinto Aduaneiro": recinto_aduaneiro_limpo,
        "Armazém": armazem_limpo,
    }


def autosize_columns(ws, max_width: int = 45):
    """
    Ajuste aproximado de largura das colunas.

    O openpyxl não executa o AutoFit real do Excel, então estimamos
    pela maior linha de texto encontrada em cada coluna e limitamos
    a largura para não deixar a planilha gigante.
    """
    for column_cells in ws.columns:
        length = 0
        col_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)

            # Se houver quebras de linha, considera a maior linha.
            parts = value.splitlines()
            longest_part = max((len(part) for part in parts), default=0)

            length = max(length, longest_part)

        ws.column_dimensions[col_letter].width = min(max(length + 2, 12), max_width)


def export_xlsx(rows, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Coordenação"

    headers = ORDERED_HEADERS
    ws.append(headers)

    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    light_border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = light_border

    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row_cells:
            cell.fill = white_fill
            cell.border = light_border
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False

    autosize_columns(ws, max_width=45)

    # Ajuda a deixar a primeira linha legível quando o cabeçalho quebra.
    ws.row_dimensions[1].height = 30

    wb.save(output_path)

def main():
    consolidated_path = latest_consolidated_json()

    with open(consolidated_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    processos = []
    for _, items in (data.get("clientes") or {}).items():
        processos.extend(items)

    rows = [build_row(p) for p in processos]

    OUTPUT_REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = OUTPUT_REPORTS_DIR / f"followup_coordenadora_{timestamp}.xlsx"

    export_xlsx(rows, output_path)

    print(f"Planilha criada em: {output_path}")
    print(f"Total de linhas: {len(rows)}")


if __name__ == "__main__":
    main()
