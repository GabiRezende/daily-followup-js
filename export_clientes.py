import json
import re
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path

try:
    from openpyxl.drawing.image import Image as XLImage
except ImportError:
    XLImage = None

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from export_coordenadora import build_row, latest_consolidated_json


OUTPUT_REPORTS_DIR = Path("output/reports")
OUTPUT_CLIENTS_DIR = Path("output/clientes")
LOGO_PATH = Path("assets/logo_js.png")

CLIENT_HEADERS = [
    "Cliente",
    "Código JS",
    "Referência Cliente",
    "Status",
    "Modal",
    "Regime",
    "Embarcação/Plataforma",
    "Incoterm",
    "Data Embarque",
    "Data Chegada",
    "Data Registro DI/DUIMP",
    "Numero DI/DUIMP",
    "Data CI",
    "Canal",
    "HAWB",
    "AWB/BL",
    "Packing List",
    "Commercial Invoice",
    "Proforma Invoice",
    "País de Origem",
]


CLIENT_DISPLAY_NAMES = {
    "ARACAJU": "ARACAJÚ",
    "SEA1": "SEA1",
    "ARACAJU_SEA1": "ARACAJÚ / SEA1",
    "CLADTEK": "CLADTEK",
    "INTERMOOR": "INTERMOOR",
    "VALUE": "VALUE",
    "FRANKS": "FRANK'S",
    "FRANK_S": "FRANK'S",
    "OCRA": "OCRA",
}


CLIENT_DELIVERY_GROUPS = {
    "ARACAJU": "ARACAJU_SEA1",
    "SEA1": "ARACAJU_SEA1",
}

CLIENT_HEADERS_BY_GROUP = {
    "ARACAJU_SEA1": [
        header for header in CLIENT_HEADERS
        if header != "Proforma Invoice"
    ],
    "CLADTEK": [
        header for header in CLIENT_HEADERS
        if header != "Embarcação/Plataforma"
    ],
    "INTERMOOR": [
        header for header in CLIENT_HEADERS
        if header != "Embarcação/Plataforma"
    ],
    "VALUE": [
        header for header in CLIENT_HEADERS
        if header != "Embarcação/Plataforma"
    ],
    "FRANK_S": [
        header for header in CLIENT_HEADERS
        if header != "Embarcação/Plataforma"
    ],
    "FRANKS": [
        header for header in CLIENT_HEADERS
        if header != "Embarcação/Plataforma"
    ],
    "OCRA": [
        header for header in CLIENT_HEADERS
        if header != "Embarcação/Plataforma"
    ],
}

def get_headers_for_client(client_key: str | None = None) -> list[str]:
    if client_key and client_key in CLIENT_HEADERS_BY_GROUP:
        return CLIENT_HEADERS_BY_GROUP[client_key]

    return CLIENT_HEADERS

def clean_text(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def strip_accents(value: str) -> str:
    text = unicodedata.normalize("NFD", value)
    return "".join(ch for ch in text if unicodedata.category(ch) != "Mn")


def normalize_for_match(value) -> str:
    text = strip_accents(clean_text(value)).upper()
    return text


def safe_client_key(value) -> str:
    text = normalize_for_match(value)
    text = re.sub(r"[^A-Z0-9]+", "_", text).strip("_")
    return text or "CLIENTE_NAO_IDENTIFICADO"


def delivery_client_key(cliente) -> str:
    client_key = safe_client_key(cliente)
    return CLIENT_DELIVERY_GROUPS.get(client_key, client_key)


def simplify_cliente(cliente) -> str:
    text = normalize_for_match(cliente)

    if "ARACAJU" in text:
        return "ARACAJU"
    if "SEA1" in text:
        return "SEA1"
    if "CLADTEK" in text:
        return "CLADTEK"
    if "PENTAGON" in text or "INTERMOOR" in text:
        return "INTERMOOR"
    if "VALUE" in text:
        return "VALUE"
    if "FRANK" in text:
        return "FRANK'S"
    if "TRAIDE" in text or "OCRA" in text:
        return "OCRA"

    return clean_text(cliente)


def extract_first(pattern: str, text: str) -> str:
    match = re.search(pattern, text, flags=re.IGNORECASE)
    if not match:
        return ""
    return clean_text(match.group(0)).upper()


def extract_all(pattern: str, text: str) -> list[str]:
    matches = re.findall(pattern, text, flags=re.IGNORECASE)
    cleaned = []

    for item in matches:
        if isinstance(item, tuple):
            item = "".join(part for part in item if part)

        item = clean_text(item).upper()

        if item and item not in cleaned:
            cleaned.append(item)

    return cleaned


def simplify_referencia(cliente: str, referencia) -> str:
    cliente = simplify_cliente(cliente)
    ref_original = clean_text(referencia)
    ref = normalize_for_match(ref_original)

    if not ref:
        return ""

    if cliente in {"ARACAJU", "SEA1"}:
        found = extract_first(r"\bSHIP\s*[-:]?\s*\d+\b", ref)
        return re.sub(r"\s*[-:]?\s*", " ", found, count=1).strip() if found else ref_original

    if cliente == "CLADTEK":
        found = []
        found.extend(extract_all(r"\bCTBI\s*[-:]?\s*\d+\b", ref))
        found.extend(extract_all(r"\bPO\s*[-:]?\s*\d+\b", ref))

        normalized = []
        for item in found:
            item = re.sub(r"\s*[-:]?\s*", " ", item, count=1).strip()

            if item not in normalized:
                normalized.append(item)

        return " / ".join(normalized) if normalized else ref_original

    if cliente == "INTERMOOR":
        found = extract_first(r"\bBRI[MA]\s*\d{3,}-\d{2,}\b", ref)
        return found.replace(" ", "") if found else ref_original

    if cliente == "VALUE":
        found = extract_first(r"\bEMB\s*\d+(?:\s+[A-Z]{3})?\b", ref)
        return found if found else ref_original

    if cliente == "OCRA":
        found = extract_first(r"\bBRI[MA]\s*\d{3,}-\d{2,}\b", ref)
        return found.replace(" ", "") if found else ref_original

    if cliente == "FRANK'S":
        found_refs = []

        # Pega FOSD, F-OSD e F-OST:
        # FOSD-000/00
        # F-OSD-056/26
        # F-OST-065/26 A
        # F-OST-065/26 B
        found_fos = extract_all(
            r"\bF[-\s]?OS[DT]\s*-?\s*\d+[A-Z]?/\d+[A-Z]?(?:\s*[AB])?\b",
            ref,
        )

        for item in found_fos:
            item = clean_text(item).upper()
            item = re.sub(r"\s*-\s*", "-", item)
            item = re.sub(r"\s*/\s*", "/", item)
            item = re.sub(r"\s+", " ", item)

            if item not in found_refs:
                found_refs.append(item)

        # Pega BRIA/BRIM:
        # BRIA1515-2026
        # BRIM1515-2026
        found_bri = extract_all(
            r"\bBRI[MA]\s*\d{3,}-\d{2,}\b",
            ref,
        )

        for item in found_bri:
            item = clean_text(item).upper().replace(" ", "")

            if item not in found_refs:
                found_refs.append(item)

        if found_refs:
            return " / ".join(found_refs)

        return ref_original

    return ref_original


def should_ignore_status(status) -> bool:
    text = normalize_for_match(status)
    return "FECHADO" in text or "EXCLUIDO" in text


def build_client_row(row: dict) -> dict:
    cliente = simplify_cliente(row.get("Cliente"))
    embarcacao = clean_text(row.get("Embarcação/Plataforma"))
    regime = clean_text(row.get("Regime"))

    if normalize_for_match(regime) == "NENHUM":
        regime = ""

    if cliente not in {"ARACAJU", "SEA1"}:
        embarcacao = ""

    return {
        "Cliente": cliente,
        "Código JS": clean_text(row.get("Código JS")),
        "Referência Cliente": simplify_referencia(cliente, row.get("Referência Cliente")),
        "Status": clean_text(row.get("Status")),
        "Modal": clean_text(row.get("Modal")),
        "Regime": regime,
        "Embarcação/Plataforma": embarcacao,
        "Incoterm": clean_text(row.get("Incoterm")),
        "Data Embarque": clean_text(row.get("Data de Embarque")),
        "Data Chegada": clean_text(row.get("Data de Chegada")),
        "Data Registro DI/DUIMP": clean_text(row.get("Data de Registro da DI")),
        "Numero DI/DUIMP": clean_text(row.get("Numero da DI")),
        "Data CI": clean_text(row.get("Data da CI")),
        "Canal": clean_text(row.get("Canal")),
        "HAWB": clean_text(row.get("House AWB/BL")),
        "AWB/BL": clean_text(row.get("Master AWB/BL")),
        "Packing List": clean_text(row.get("Pack List")),
        "Commercial Invoice": clean_text(row.get("Commercial Invoice")),
        "Proforma Invoice": clean_text(row.get("Proforma Invoice")),
        "País de Origem": clean_text(row.get("País de Origem")),
    }


def autosize_columns(ws, max_width: int = 42):
    for column_cells in ws.columns:
        length = 0
        col_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)

            # Se tiver quebra de linha, considera a maior linha,
            # não o texto inteiro somado.
            parts = value.splitlines()
            longest_part = max((len(part) for part in parts), default=0)

            length = max(length, longest_part)

        ws.column_dimensions[col_letter].width = min(max(length + 2, 12), max_width)


def apply_preferred_widths(ws, header_row: int, headers: list[str]):
    preferred_widths = {
        "Cliente": 12,
        "Código JS": 12,
        "Referência Cliente": 36,
        "Status": 32,
        "Modal": 18,
        "Regime": 18,
        "Embarcação/Plataforma": 24,
        "Incoterm": 14,
        "Data Embarque": 20,
        "Data Chegada": 20,
        "Data Registro DI/DUIMP": 28,
        "Numero DI/DUIMP": 28,
        "Data CI": 16,
        "Canal": 14,
        "HAWB": 20,
        "AWB/BL": 22,
        "Packing List": 34,
        "Commercial Invoice": 30,
        "Proforma Invoice": 30,
        "País de Origem": 20,
    }

    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        preferred = preferred_widths.get(header)

        if preferred:
            current = ws.column_dimensions[col_letter].width or 0
            ws.column_dimensions[col_letter].width = max(current, preferred)

def sanitize_sheet_title(value, fallback: str = "Follow-up") -> str:
    title = clean_text(value) or fallback

    # Excel não aceita estes caracteres em nome de aba: \ / * ? : [ ]
    title = re.sub(r'[\\/*?:\[\]]', "-", title)

    title = re.sub(r"\s+", " ", title).strip()
    title = title[:31]

    return title or fallback

def add_logo(ws, logo_path: Path):
    if not logo_path.exists():
        return

    if XLImage is None:
        print("Logo não inserida: instale pillow com 'python -m pip install pillow'")
        return

    try:
        img = XLImage(str(logo_path))
        img.width = 145
        img.height = 55
        ws.add_image(img, "A1")
    except Exception as exc:
        print(f"Logo não inserida: {exc}")
        

def export_xlsx(
    rows: list[dict],
    output_path: Path,
    sheet_title: str = "Acompanhamento Diário",
    report_date: str | None = None,
    headers: list[str] | None = None,
):
    headers = headers or CLIENT_HEADERS

    wb = Workbook()
    ws = wb.active
    ws.title = sanitize_sheet_title(sheet_title)

    report_date = report_date or datetime.now().strftime("%d/%m/%Y")
    last_col = get_column_letter(len(headers))
    header_row = 5
    first_data_row = header_row + 1

    add_logo(ws, LOGO_PATH)

    # Título principal
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=3)
    title_cell.value = f"Acompanhamento Diário dos Processos em Aberto - {sheet_title}"
    title_cell.font = Font(bold=True, size=16, color="1F4E78")
    title_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    # Data
    ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=len(headers))
    date_cell = ws.cell(row=2, column=3)
    date_cell.value = f"Atualizado em: {report_date}"
    date_cell.font = Font(size=11, color="404040")
    date_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    # Subtítulo
    ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=len(headers))
    subtitle_cell = ws.cell(row=3, column=3)
    subtitle_cell.value = ""
    subtitle_cell.font = Font(size=10, italic=True, color="666666")
    subtitle_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    # Ajuste do topo
    ws.row_dimensions[1].height = 42
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 8

    # Cabeçalho
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = header

    # Dados
    for row_idx, row_data in enumerate(rows, start=first_data_row):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx).value = row_data.get(header, "")

    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(color="FFFFFF", bold=True)

    white_fill = PatternFill("solid", fgColor="FFFFFF")
    light_border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    # Colunas que podem quebrar linha porque costumam ter texto longo
    wrap_columns = {
        "Referência Cliente",
        "Status",
        "Regime",
        "Embarcação/Plataforma",
        "Packing List",
        "Commercial Invoice",
        "Proforma Invoice",
        "País de Origem",
    }

    for row_cells in ws.iter_rows(
        min_row=header_row,
        max_row=ws.max_row,
        min_col=1,
        max_col=len(headers),
    ):
        for cell in row_cells:
            cell.border = light_border

            header = headers[cell.column - 1]

            if cell.row == header_row:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
            else:
                cell.fill = white_fill
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=header in wrap_columns,
                )

    ws.freeze_panes = f"A{first_data_row}"
    ws.auto_filter.ref = f"A{header_row}:{last_col}{ws.max_row}"
    ws.sheet_view.showGridLines = False

    autosize_columns(ws, max_width=42)
    apply_preferred_widths(ws, header_row, headers)

    # Espaço da logo
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 12, 16)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 12, 16)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def load_processes_from_latest_json() -> list[dict]:
    consolidated_path = latest_consolidated_json()

    with open(consolidated_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    processos = []
    for _, items in (data.get("clientes") or {}).items():
        processos.extend(items)

    return processos


def main():
    processos = load_processes_from_latest_json()
    base_rows = [build_row(p) for p in processos]

    client_rows = []

    for row in base_rows:
        if should_ignore_status(row.get("Status")):
            continue

        client_rows.append(build_client_row(row))

    client_rows.sort(
        key=lambda x: (
            delivery_client_key(x.get("Cliente")),
            x.get("Cliente", ""),
            x.get("Código JS", ""),
        )
    )

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    run_date = datetime.now().strftime("%d/%m/%Y")

    OUTPUT_REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    general_path = OUTPUT_REPORTS_DIR / f"followup_clientes_{timestamp}.xlsx"
    export_xlsx(
        client_rows,
        general_path,
        sheet_title="Clientes",
        report_date=run_date,
        headers=CLIENT_HEADERS,
    )

    grouped = defaultdict(list)

    for row in client_rows:
        grouped[delivery_client_key(row.get("Cliente"))].append(row)

    run_dir = OUTPUT_CLIENTS_DIR / timestamp
    run_dir.mkdir(parents=True, exist_ok=True)

    manifest = {
        "generated_at": timestamp,
        "date_br": run_date,
        "general_file": str(general_path),
        "folder": str(run_dir),
        "files": [],
    }

    for client_key, rows in sorted(grouped.items()):
        display_name = CLIENT_DISPLAY_NAMES.get(
            client_key,
            rows[0].get("Cliente") or client_key,
        )

        file_path = run_dir / f"{client_key}_followup_{timestamp}.xlsx"
        export_xlsx(
            rows,
            file_path,
            sheet_title=display_name,
            report_date=run_date,
            headers=get_headers_for_client(client_key),
        )

        manifest["files"].append({
            "client_key": client_key,
            "client": display_name,
            "path": str(file_path),
            "row_count": len(rows),
        })

    manifest_path = run_dir / "manifest.json"
    manifest_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(f"Planilha geral de clientes criada em: {general_path}")
    print(f"Arquivos individuais criados em: {run_dir}")
    print(f"Manifest criado em: {manifest_path}")
    print(f"Total de linhas para clientes: {len(client_rows)}")
    print(f"Total de clientes/arquivos: {len(manifest['files'])}")


if __name__ == "__main__":
    main()